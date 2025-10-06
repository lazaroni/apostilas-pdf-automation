#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import io
import os
import random
import re
import string
import sys
import winsound
from pathlib import Path

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.colors import Color
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pypdf import PdfReader, PdfWriter
from pypdf.generic import RectangleObject, NameObject
from tqdm import tqdm

# ================================================
# Configurações para buscar arquivos na pasta raiz
# ================================================

RAIZ = Path.cwd()

# =========================
# Configurações ajustáveis
# =========================

FONT_SIZE_PT = 12
TEXT_COLOR = None  # será definida dinamicamente no main()
STAMP_ALL_PAGES = True
OUTPUT_DIR_NAME = "saida_pdfs"

ARIAL_CANDIDATES = [
    "Arial.ttf",
    "arial.ttf",
    "C:/Windows/Fonts/arial.ttf",
    "/Library/Fonts/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf",
    "/usr/share/fonts/truetype/msttcorefonts/arial.ttf",
]

# =========================
# Funções utilitárias
# =========================

def cm_to_pt(cm: float) -> float:
    return (cm / 2.54) * 72.0

def register_arial_or_fallback() -> str:
    for candidate in ARIAL_CANDIDATES:
        if os.path.isfile(candidate):
            try:
                pdfmetrics.registerFont(TTFont("Arial", candidate))
                print(f"[Fonte] Arial registrada com sucesso: {candidate}")
                return "Arial"
            except Exception:
                pass
    print("[Fonte] Arial não encontrada. Usando Helvetica como fallback.")
    return "Helvetica"

def normalize_cpf(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    digits = re.sub(r"\D", "", s)
    return digits if digits else None

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\/\\\:\*\?\"\<\>\|]', "_", name).strip()

def make_overlay(page_width: float, page_height: float, text: str, font_name: str) -> PdfReader:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(page_width, page_height))
    c.setFillColor(TEXT_COLOR)
    c.setFont(font_name, FONT_SIZE_PT)
    x = cm_to_pt(1.5)
    y = page_height - cm_to_pt(0.5) - FONT_SIZE_PT
    c.drawString(x, y, text)
    c.save()
    buf.seek(0)
    return PdfReader(buf)

def normalizar_pdf_temporario(entrada: Path) -> Path:
    print(f"[PDF] Normalizando PDF: {entrada.name}")
    leitor = PdfReader(str(entrada))
    gravador = PdfWriter()

    for pagina in leitor.pages:
        obj = pagina.get_object()
        if "/MediaBox" not in obj:
            obj[NameObject("/MediaBox")] = RectangleObject([0, 0, 595, 842])
        gravador.add_page(pagina)

    gravador.remove_links()
    gravador.add_metadata({})
    gravador.encrypt("")

    saida = RAIZ / "modelo-normalizado.pdf"
    with open(saida, "wb") as f:
        gravador.write(f)

    print(f"[PDF] PDF normalizado salvo como: {saida.name}")
    return saida

def read_students_from_xlsx(xlsx_path: Path) -> list[tuple[str, str]]:
    print(f"[Planilha] Lendo dados de: {xlsx_path.name}")
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb.active
    students: list[tuple[str, str]] = []

    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        name_val, cpf_val = row
        if name_val is None:
            continue
        name = str(name_val).strip()
        cpf = normalize_cpf(cpf_val)
        if not name or not cpf:
            continue
        students.append((name.upper(), cpf))

    print(f"[Planilha] {len(students)} alunos encontrados.")
    return students

def tocar_som_assincrono(nome_arquivo):
    try:
        winsound.PlaySound(nome_arquivo, winsound.SND_FILENAME | winsound.SND_ASYNC)
    except Exception as e:
        print(f"⚠️ Erro ao tocar {nome_arquivo}: {e}")

# =========================
# Funções de geração
# =========================

def stamp_pdf_for_student(src_pdf_path: Path, out_pdf_path: Path, text_line: str, font_name: str):
    reader = PdfReader(str(src_pdf_path))
    writer = PdfWriter()
    page_bar = tqdm(total=len(reader.pages), desc="📄 Páginas", unit="página", leave=True)

    for i, page in enumerate(reader.pages):
        if (i == 0) or STAMP_ALL_PAGES:
            media_box = page.mediabox
            pw = float(media_box.width)
            ph = float(media_box.height)
            overlay_reader = make_overlay(pw, ph, text_line, font_name)
            overlay_page = overlay_reader.pages[0]
            page.merge_page(overlay_page)

        writer.add_page(page)
        page_bar.update(1)

    page_bar.close()

    writer.add_metadata({
        "/Producer": "Python pypdf",
        "/Author": "Automação",
        "/Title": text_line
    })

    out_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_pdf_path, "wb") as f:
        writer.write(f)

    print(f"[✔] PDF salvo: {out_pdf_path.name}")

def gerar_pdfs_coringa(pdf_path: Path, sigla: str, text_color: Color):
    try:
        quantidade = int(input("\nQuantas apostilas coringa pretendes gerar? ").strip())
        if quantidade <= 0:
            raise ValueError
    except ValueError:
        print("[Erro] Valor inválido. Digite um número inteiro positivo.")
        sys.exit(1)

    global TEXT_COLOR
    TEXT_COLOR = text_color

    font_name = register_arial_or_fallback()
    out_dir = pdf_path.parent / "saida_coringa"
    out_dir.mkdir(exist_ok=True)

    print(f"\n📦 Gerando {quantidade} PDFs coringa...\n")

    for i in range(quantidade):
        texto_original = gerar_texto_coringa(sigla)
        texto_com_espacos = texto_original.replace("_", " ")

        file_name = sanitize_filename(texto_com_espacos) + ".pdf"
        out_path = out_dir / file_name

        stamp_pdf_coringa(pdf_path, out_path, texto_com_espacos, font_name)

        if i < quantidade - 1:
            tocar_som_assincrono("whatsapp_message.wav")

    tocar_som_assincrono("wow.wav")
    print(f"\n✅ Concluído. Arquivos gerados em: {out_dir}\n")

    try:
        os.startfile(out_dir)
    except Exception as e:
        print(f"\n⚠️ Erro ao abrir o diretório: {e}")

def gerar_pdfs_com_planilha(pdf_path: Path, xlsx_path: Path, sigla: str, text_color: Color):
    global TEXT_COLOR
    TEXT_COLOR = text_color

    students = read_students_from_xlsx(xlsx_path)
    if not students:
        print("[Erro] Nenhum aluno válido encontrado na planilha.")
        sys.exit(1)

    font_name = register_arial_or_fallback()
    out_dir = pdf_path.parent / OUTPUT_DIR_NAME
    out_dir.mkdir(exist_ok=True)

    print(f"\n📦 Gerando PDFs para {len(students)} alunos...\n")

    for i, (name, cpf) in enumerate(students):
        text_line = f"{sigla} {name} CPF {cpf}"
        file_name = sanitize_filename(text_line) + ".pdf"
        out_path = out_dir / file_name
        stamp_pdf_for_student(pdf_path, out_path, text_line, font_name)

        if i < len(students) - 1:
            tocar_som_assincrono("whatsapp_message.wav")

    tocar_som_assincrono("wow.wav")
    print(f"\n✅ Concluído. Arquivos gerados em: {out_dir}\n")

    try:
        os.startfile(out_dir)
    except Exception as e:
        print(f"\n⚠️ Erro ao abrir o diretório: {e}")

def gerar_codigo(prefixo_letras: int, sufixo_numeros: int) -> str:
    letras = ''.join(random.choices(string.ascii_uppercase, k=prefixo_letras))
    numeros = ''.join(random.choices(string.digits, k=sufixo_numeros))
    return letras + numeros

def gerar_texto_coringa(sigla: str) -> str:
    localizador = gerar_codigo(3, 3)
    aluno = gerar_codigo(2, 3)
    return f"{sigla}_LOCALIZADOR_{localizador}_ALUNO_{aluno}_PROIBIDA_A_DIVULGAÇÃO_DESSE_PDF"

def stamp_pdf_coringa(src_pdf_path: Path, out_pdf_path: Path, text_line: str, font_name: str):
    reader = PdfReader(str(src_pdf_path))
    writer = PdfWriter()
    page_bar = tqdm(total=len(reader.pages), desc="📄 Páginas", unit="página", leave=True)

    for i, page in enumerate(reader.pages):
        if (i == 0) or STAMP_ALL_PAGES:
            media_box = page.mediabox
            pw = float(media_box.width)
            ph = float(media_box.height)
            overlay_reader = make_overlay(pw, ph, text_line, font_name)
            overlay_page = overlay_reader.pages[0]
            page.merge_page(overlay_page)

        writer.add_page(page)
        page_bar.update(1)

    page_bar.close()

    writer.add_metadata({
        "/Producer": "Python pypdf",
        "/Author": "Registro Coringa",
        "/Title": text_line
    })

    out_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_pdf_path, "wb") as f:
        writer.write(f)
        
    print(f"[✔] PDF salvo: {out_pdf_path.name}")

def main():
    print("📄 Iniciando processo de geração de PDFs personalizados...\n")
    print("📁 Certifique-se de que os arquivos 'modelo.pdf' e 'dados.xlsx' estão na mesma pasta deste script.\n")

    # Entrada da sigla personalizada
    sigla = input("Informe a sigla utilizada pela apostila personalizada: ").strip().upper()

    # Escolha da cor do texto
    print("\nEscolha a cor do texto estampado no PDF:")
    print("1. Padrão Vermelho Claro")
    print("2. Cor personalizada")
    opcao_cor = input("Digite 1 ou 2: ").strip()

    global TEXT_COLOR
    if opcao_cor == "1":
        TEXT_COLOR = Color(1, 0, 0)
    elif opcao_cor == "2":
        hex_input = input("Digite o código hexadecimal da cor (ex: #00FF00): ").strip().lstrip("#")
        try:
            r = int(hex_input[0:2], 16) / 255
            g = int(hex_input[2:4], 16) / 255
            b = int(hex_input[4:6], 16) / 255
            TEXT_COLOR = Color(r, g, b)
        except Exception:
            print("[Erro] Código hexadecimal inválido. Usando vermelho claro como fallback.")
            TEXT_COLOR = Color(1, 0, 0)
    else:
        print("[Erro] Opção inválida. Usando vermelho claro como fallback.")
        TEXT_COLOR = Color(1, 0, 0)

    # Caminhos padrão
    CAMINHO_PADRAO_PDF = RAIZ / "modelo.pdf"
    CAMINHO_PADRAO_XLSX = RAIZ / "dados.xlsx"

    # Compatibilidade do PDF
    print("\nVocê já possui um arquivo PDF compatível?")
    print("1. Sim")
    print("2. Não")
    opcao_pdf_compativel = input("Digite 1 ou 2: ").strip()

    # Seleção do PDF-base
    print("\nSelecione o arquivo PDF:")
    print("1. Arquivo padrão (modelo.pdf)")
    print("2. Arquivo personalizado ou de fora da pasta raiz")
    opcao_pdf = input("Digite 1 ou 2: ").strip()

    if opcao_pdf == "1":
        pdf_in = CAMINHO_PADRAO_PDF
    else:
        pdf_in = input("Caminho do PDF-base: ").strip().strip('"').strip("'")

    pdf_original = Path(pdf_in).expanduser().resolve()
    if not pdf_original.exists():
        print(f"[Erro] Arquivo PDF não encontrado: {pdf_original}")
        sys.exit(1)

    if opcao_pdf_compativel == "1":
        pdf_path = pdf_original
        print(f"[PDF] Usando PDF original sem normalização: {pdf_path.name}")
    elif opcao_pdf_compativel == "2":
        pdf_path = normalizar_pdf_temporario(pdf_original)
    else:
        print("[Erro] Opção inválida. Digite apenas 1 ou 2.")
        sys.exit(1)

    # Escolha do tipo de geração
    print("\nPrecisa gerar:")
    print("1. PDFs utilizando dados de uma planilha")
    print("2. PDFs coringa")
    opcao_tipo_pdf = input("Digite 1 ou 2: ").strip()

    if opcao_tipo_pdf == "1":
        print("\nSelecione a planilha XLSX:")
        print("1. Arquivo padrão (dados.xlsx)")
        print("2. Arquivo personalizado ou de fora da pasta raiz")
        opcao_xlsx = input("Digite 1 ou 2: ").strip()

        if opcao_xlsx == "1":
            xlsx_in = CAMINHO_PADRAO_XLSX
        else:
            xlsx_in = input("Caminho da planilha XLSX: ").strip().strip('"').strip("'")

        xlsx_path = Path(xlsx_in).expanduser().resolve()
        if not xlsx_path.exists():
            print(f"[Erro] Planilha XLSX não encontrada: {xlsx_path}")
            sys.exit(1)

        gerar_pdfs_com_planilha(pdf_path, xlsx_path, sigla, TEXT_COLOR)

    elif opcao_tipo_pdf == "2":
        gerar_pdfs_coringa(pdf_path, sigla, TEXT_COLOR)

    else:
        print("[Erro] Opção inválida. Digite apenas 1 ou 2.")
        sys.exit(1)

if __name__ == "__main__":
    main()
